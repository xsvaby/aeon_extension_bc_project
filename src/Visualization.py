from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

import requests
from typing import Dict, List, Set, Tuple, Any, Iterable
from SPARQLWrapper import JSON
import networkx as nx
import matplotlib.pyplot as plt

from EnrichmentClasses import EnrichmentGOterm, EnrichmentPSBN


def append_column_to_xlsx(filepath: str, data: Iterable[Any], column_name: str = "Col") -> None:
    """
    Append data as a new column to an existing XLSX file.

    If the file does not exist, it will be created. The column name is written
    to the first row, and the data values are written starting from row 2.

    Parameters
    ----------
    filepath : str
        Path to the XLSX file.
    data : Iterable[Any]
        Iterable of values to insert as a column.
    column_name : str, optional
        Name of the column header (default is "Col").

    Returns
    -------
    None
    """
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    if ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
        col_index: int = 1
    else:
        col_index = ws.max_column + 1

    col_letter: str = get_column_letter(col_index)
    ws[f"{col_letter}1"] = column_name

    for i, value in enumerate(data, start=2):
        ws[f"{col_letter}{i}"] = value

    wb.save(filepath)


def get_quickgo_terms_batch(go_ids: Set[str]) -> List[Dict[str, Any]]:
    """
    Fetch GO term metadata from the QuickGO API in batch mode.

    Parameters
    ----------
    go_ids : set[str]
        Set of GO term IDs to query.

    Returns
    -------
    list[dict]
        List of GO term records returned by the QuickGO API.
    """
    QUICKGO_BATCH_URL: str = "https://www.ebi.ac.uk/QuickGO/services/ontology/go/terms/{}"
    joined: str = ",".join(go_ids)
    url: str = QUICKGO_BATCH_URL.format(joined)

    r = requests.get(url, headers={"Accept": "application/json"})
    r.raise_for_status()

    data: Dict[str, Any] = r.json()
    return data.get("results", [])


def set_nodes_for_graph(intersected_goterms: Dict[str, EnrichmentGOterm]) -> None:
    """
    Populate parent/child relationships among intersected GO terms
    using data retrieved from the QuickGO API.

    This function modifies `intersected_goterms` in-place by assigning:
    - parents to child nodes
    - children (with relations) to parent nodes

    Parameters
    ----------
    intersected_goterms : dict[str, EnrichmentGOterm]
        Mapping of GO IDs to GO term objects that are part of the intersection.
    """
    intersected_go_ids: Set[str] = set(intersected_goterms.keys())
    terms: List[Dict[str, Any]] = get_quickgo_terms_batch(intersected_go_ids)

    for term in terms:
        current_term_id: str = term["id"]
        children: List[Dict[str, Any]] = term.get("children", [])

        for child in children:
            child_id: str = child.get("id", "")
            child_relation: str = child.get("relation", "")

            if child_id in intersected_go_ids:
                intersected_goterms[child_id].add_parent(
                    intersected_goterms[current_term_id]
                )
                intersected_goterms[current_term_id].add_child(
                    intersected_goterms[child_id], child_relation
                )


def make_graph(intersected_goterms: Dict[str, EnrichmentGOterm]) -> nx.DiGraph:
    """
    Build a directed graph representing GO term parent-child relationships.

    Parameters
    ----------
    intersected_goterms : dict[str, EnrichmentGOterm]
        Mapping of GO IDs to GO term objects.

    Returns
    -------
    networkx.DiGraph
        Directed graph where nodes are GO terms and edges represent
        parent-child relationships.
    """
    G: nx.DiGraph = nx.DiGraph()

    for term_id, term in intersected_goterms.items():
        G.add_node(term_id, label=getattr(term, "name", term_id))

        for child_term, relation in term.children.items():
            child_id: str = child_term.go_id
            G.add_node(child_id, label=getattr(child_term, "name", child_id))
            G.add_edge(term_id, child_id, relation=relation)

    return G


def visualize_subgraphs(
    G: nx.DiGraph,
    sorted_roots: List[EnrichmentGOterm],
    intersected_goterms: Dict[str, EnrichmentGOterm],
) -> None:
    """
    Visualize subgraphs of a GO term graph rooted at each root node.

    Each root generates a separate plotted subgraph.

    Parameters
    ----------
    G : networkx.DiGraph
        Full GO term graph.
    sorted_roots : list[EnrichmentGOterm]
        List of root GO terms sorted by FDR.
    intersected_goterms : dict[str, EnrichmentGOterm]
        Mapping of GO IDs to GO term objects.
    """
    for root_term in sorted_roots:
        root_id: str = root_term.go_id

        # Build subgraph
        desc: Set[str] = nx.descendants(G, root_id)
        subnodes: Set[str] = desc | {root_id}
        H: nx.DiGraph = G.subgraph(subnodes).copy()

        # Draw subgraph
        plt.figure(figsize=(18, 14))

        pos = nx.nx_agraph.graphviz_layout(H, prog="dot")

        nx.draw_networkx_nodes(H, pos, node_size=800)
        nx.draw_networkx_edges(H, pos, arrowstyle="<-", arrowsize=20, width=2)

        node_labels: Dict[str, str] = {
            n: intersected_goterms[n].process_name for n in H.nodes
        }
        nx.draw_networkx_labels(H, pos, labels=node_labels, font_size=7)

        edge_labels: Dict[Tuple[str, str], str] = nx.get_edge_attributes(H, "relation")
        nx.draw_networkx_edge_labels(H, pos, edge_labels=edge_labels, font_size=7)

        plt.title(f"Subgraph rooted at {root_term.process_name}")
        plt.axis("off")
        plt.tight_layout()
        plt.show()


def get_roots_and_leafs(
    intersected_goterms: Dict[str, EnrichmentGOterm]
) -> Tuple[Set[EnrichmentGOterm], Set[EnrichmentGOterm]]:
    """
    Identify root and leaf GO terms from a GO term graph.

    Roots are terms without parents.
    Leafs are terms without children.

    Parameters
    ----------
    intersected_goterms : dict[str, EnrichmentGOterm]
        Mapping of GO IDs to GO term objects.

    Returns
    -------
    tuple[set[EnrichmentGOterm], set[EnrichmentGOterm]]
        A tuple containing:
        - Set of root GO terms
        - Set of leaf GO terms
    """
    roots: Set[EnrichmentGOterm] = set()
    leafs: Set[EnrichmentGOterm] = set()

    for goterm in intersected_goterms.values():
        if goterm.parents == {}:
            roots.add(goterm)
        if goterm.children == {}:
            leafs.add(goterm)

    return roots, leafs


def sort_roots_and_leafs(
    roots: Set[EnrichmentGOterm],
    leafs: Set[EnrichmentGOterm],
) -> Tuple[List[EnrichmentGOterm], List[EnrichmentGOterm]]:
    """
    Sort root and leaf GO terms by FDR value.

    Parameters
    ----------
    roots : set[EnrichmentGOterm]
        Set of root GO terms.
    leafs : set[EnrichmentGOterm]
        Set of leaf GO terms.

    Returns
    -------
    tuple[list[EnrichmentGOterm], list[EnrichmentGOterm]]
        Sorted lists of root and leaf GO terms.
    """
    sorted_roots: List[EnrichmentGOterm] = sorted(roots, key=lambda t: t.fdr)
    sorted_leafs: List[EnrichmentGOterm] = sorted(leafs, key=lambda t: t.fdr)

    return sorted_roots, sorted_leafs


def visualize_subgraphs_on_whole_net(psbn: EnrichmentPSBN) -> None:
    """
    Visualize GO term subgraphs across the entire PSBN network.

    Parameters
    ----------
    psbn : EnrichmentPSBN
        PSBN object containing enrichment analysis for multiple instances.
    """
    intersected_goterms = psbn.goterms_intersection_on_all_instances()
    set_nodes_for_graph(intersected_goterms)

    roots, leafs = get_roots_and_leafs(intersected_goterms)
    sorted_roots, _ = sort_roots_and_leafs(roots, leafs)

    G = make_graph(intersected_goterms)
    visualize_subgraphs(G, sorted_roots, intersected_goterms)


def print_roots_and_leafs_on_whole_net(psbn: EnrichmentPSBN) -> None:
    """
    Print root and leaf GO terms for the entire PSBN network.

    Parameters
    ----------
    psbn : EnrichmentPSBN
        PSBN object containing enrichment analysis.
    """
    goterm_intersection = psbn.goterms_intersection_on_all_instances()
    roots, leafs = get_roots_and_leafs(goterm_intersection)
    sorted_roots, sorted_leafs = sort_roots_and_leafs(roots, leafs)

    print(f"{len(sorted_leafs)} leafs: {sorted_leafs}")
    print(f"{len(sorted_roots)} roots: {sorted_roots}")


def visualize_subgraphs_on_each_instance(psbn: EnrichmentPSBN) -> None:
    """
    Visualize GO term subgraphs for each PSBN instance separately.

    Parameters
    ----------
    psbn : EnrichmentPSBN
        PSBN object containing multiple enrichment instances.
    """
    for i, psbn_instance in enumerate(psbn.instances):
        print(f"{i}: {psbn_instance.color}")

        intersected_goterms = psbn_instance.goterm_intersection()
        set_nodes_for_graph(intersected_goterms)

        roots, leafs = get_roots_and_leafs(intersected_goterms)
        sorted_roots, _ = sort_roots_and_leafs(roots, leafs)

        G = make_graph(intersected_goterms)
        visualize_subgraphs(G, sorted_roots, intersected_goterms)


def print_roots_and_leafs_per_instance(psbn: EnrichmentPSBN) -> None:
    """
    Print root and leaf GO terms separately for each PSBN instance.

    Parameters
    ----------
    psbn : EnrichmentPSBN
        PSBN object containing multiple enrichment instances.
    """
    for psbn_instance in psbn.instances:
        goterm_intersection = psbn_instance.goterm_intersection()

        print(psbn_instance.color)

        roots, leafs = get_roots_and_leafs(goterm_intersection)
        sorted_roots, sorted_leafs = sort_roots_and_leafs(roots, leafs)

        print(f"{len(sorted_leafs)} leafs: {sorted_leafs}")
        print(f"{len(sorted_roots)} roots: {sorted_roots}")
